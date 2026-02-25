from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from pydantic import BaseModel
import csv
import io
from datetime import datetime
import sys
from pathlib import Path
import openpyxl
import xlrd
import json
import itertools
import re

# Add backend directory to path for imports
sys.path.insert(0, str(Path(__file__).parent))
from services.rule_engine import apply_rule

# Database connection
DATABASE_URL = "postgresql://postgres:AditiRao@127.0.0.1/mdm_db"
engine = create_engine(DATABASE_URL)
SessionLocal = sessionmaker(bind=engine)

app = FastAPI(title="MDM-Dev System")

# Add CORS middleware to allow frontend to communicate with backend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://localhost:3001", "http://localhost:3002", "http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def auto_migrate_database():
    """
    Automatically create missing columns if they don't exist.
    This runs on startup to ensure schema is ready.
    """
    try:
        with engine.connect() as conn:
            # Add columns_info to jobs table
            try:
                conn.execute(text("""
                    ALTER TABLE jobs ADD COLUMN IF NOT EXISTS columns_info TEXT;
                """))
                conn.commit()
                print("✓ Ensured jobs.columns_info column exists")
            except Exception as e:
                print(f"Note: jobs.columns_info check: {str(e)}")
            
            # Add row_data to clean_data table
            try:
                conn.execute(text("""
                    ALTER TABLE clean_data ADD COLUMN IF NOT EXISTS row_data TEXT;
                """))
                conn.commit()
                print("✓ Ensured clean_data.row_data column exists")
            except Exception as e:
                print(f"Note: clean_data.row_data check: {str(e)}")
            
            # Add row_data to quarantine_data table
            try:
                conn.execute(text("""
                    ALTER TABLE quarantine_data ADD COLUMN IF NOT EXISTS row_data TEXT;
                """))
                conn.commit()
                print("✓ Ensured quarantine_data.row_data column exists")
            except Exception as e:
                print(f"Note: quarantine_data.row_data check: {str(e)}")
    except Exception as e:
        print(f"Auto-migration warning: {str(e)}")

# Run auto-migration on startup
auto_migrate_database()

@app.get("/")
def read_root():
    """Health check endpoint"""
    return {"message": "MDM-Dev System is running"}

@app.post("/setup-rules")
def setup_rules():
    """
    Initialize sample rules - SETUP ENDPOINT
    """
    try:
        with engine.connect() as conn:
            # Delete existing rules first
            conn.execute(text("DELETE FROM rules"))
            
            # Insert sample rules
            conn.execute(text("""
                INSERT INTO rules (column_name, rule_type, rule_value, is_active)
                VALUES 
                    ('name', 'regex', '^[A-Za-z ]+$', TRUE),
                    ('age', 'range', '0-120', TRUE)
            """))
            conn.commit()
        return {"message": "Rules initialized successfully"}
    except Exception as e:
        return {"error": str(e)}

@app.post("/update-rule")
def update_rule(column_name: str, rule_type: str, rule_value: str):
    """Update a rule in the database"""
    try:
        with engine.connect() as conn:
            conn.execute(text("""
                UPDATE rules
                SET rule_type = :type, rule_value = :value
                WHERE column_name = :column
            """), {
                "type": rule_type,
                "value": rule_value,
                "column": column_name
            })
            conn.commit()
        return {"message": "Rule updated successfully", "column": column_name, "rule_value": rule_value}
    except Exception as e:
        return {"error": str(e)}


class RuleUpdate(BaseModel):
    column_name: str
    rule_type: str
    rule_value: str

@app.put("/rules/{rule_id}")
def update_rule_put(rule_id: int, rule: RuleUpdate):
    """Update a rule via PUT request with JSON body"""
    try:
        with engine.connect() as conn:
            conn.execute(text("""
                UPDATE rules
                SET column_name = :column, rule_type = :type, rule_value = :value
                WHERE id = :id
            """), {
                "id": rule_id,
                "column": rule.column_name,
                "type": rule.rule_type,
                "value": rule.rule_value
            })
            conn.commit()
        return {"message": "Rule updated successfully"}
    except Exception as e:
        return {"error": str(e)}

@app.post("/preview-file")
async def preview_file(file: UploadFile = File(...)):
    """
    Preview file headers and first few rows without processing the entire file.
    Detects column data types and checks for existing rules.
    Returns metadata for rule configuration.
    """
    try:
        filename_lower = file.filename.lower()
        contents = await file.read()
        headers = []
        sample_rows = []
        
        # Parse file based on extension
        if filename_lower.endswith('.csv'):
            csv_reader = csv.DictReader(io.StringIO(contents.decode('utf-8')))
            if csv_reader.fieldnames is None:
                raise HTTPException(status_code=400, detail="CSV file is empty or invalid")
            headers = list(csv_reader.fieldnames)
            sample_rows = list(itertools.islice(csv_reader, 5))  # Get first 5 rows
        
        elif filename_lower.endswith('.xlsx'):
            excel_book = openpyxl.load_workbook(io.BytesIO(contents))
            sheet = excel_book.active
            headers = [cell.value for cell in sheet[1]]
            if not headers or headers == [None]:
                raise HTTPException(status_code=400, detail="Excel file is empty or has no headers")
            for row in itertools.islice(sheet.iter_rows(min_row=2, values_only=True), 5):
                if all(cell is None for cell in row):
                    continue
                sample_rows.append(dict(zip(headers, row)))
        
        elif filename_lower.endswith('.xls'):
            excel_book = xlrd.open_workbook(file_contents=contents)
            sheet = excel_book.sheet_by_index(0)
            if sheet.nrows == 0:
                raise HTTPException(status_code=400, detail="XLS file is empty")
            headers = [str(sheet.cell_value(0, col_idx)) for col_idx in range(sheet.ncols)]
            for row_idx in range(1, min(6, sheet.nrows)):
                row_values = [sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)]
                if all(not val for val in row_values):
                    continue
                sample_rows.append(dict(zip(headers, row_values)))
        
        else:
            raise HTTPException(status_code=400, detail="File must be CSV, XLS, or XLSX format")
        
        # Fetch system rules from database
        with engine.connect() as conn:
            rules_result = conn.execute(text("""
                SELECT column_name, rule_type, rule_value
                FROM rules
                WHERE is_active = TRUE
            """))
            system_rules = {r[0]: {"type": r[1], "value": r[2]} for r in rules_result.fetchall()}
        
        # Auto-detect column data types
        column_metadata = []
        for header in headers:
            detected_type = "text"  # default
            
            # Sample data from first few rows
            if sample_rows:
                sample_values = [str(row.get(header, "")).strip() for row in sample_rows if header in row]
                
                # Detect type based on sample values
                if sample_values:
                    # Check if all numeric
                    try:
                        all(float(v) for v in sample_values if v)
                        detected_type = "number"
                    except:
                        pass
                    
                    # Check if email-like
                    if detected_type == "text" and all("@" in v for v in sample_values if v):
                        detected_type = "email"
                    
                    # Check if date-like
                    if detected_type == "text" and sample_values:
                        date_patterns = [r'\d{1,4}[-/]\d{1,2}[-/]\d{1,4}', r'\d{4}-\d{2}-\d{2}']
                        if any(any(re.search(pattern, v) for pattern in date_patterns) for v in sample_values if v):
                            detected_type = "date"
            
            system_rule = system_rules.get(header, None)
            
            column_metadata.append({
                "name": header,
                "detected_type": detected_type,
                "system_rule": system_rule,
                "custom_rule": None  # Will be set by user
            })
        
        return {
            "file_name": file.filename,
            "columns": column_metadata,
            "sample_rows": sample_rows,
            "total_columns": len(headers)
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/upload")
async def upload_csv(file: UploadFile = File(...), column_rules: str = None):
    """
    Upload CSV or Excel file for data quality validation.
    Applies rules from database to validate each row.
    Stores all columns dynamically.

    Supports: .csv, .xls, .xlsx
    
    Optional: column_rules - JSON string with custom rules for each column
    Format: {"column_name": {"type": "regex", "value": "pattern"}, ...}
    """
    try:
        # Validate file type
        filename_lower = file.filename.lower()
        if not (filename_lower.endswith('.csv') or filename_lower.endswith('.xls') or filename_lower.endswith('.xlsx')):
            raise HTTPException(status_code=400, detail="File must be CSV, XLS, or XLSX format")
        
        # Read file contents
        contents = await file.read()
        rows_list = []
        columns = []
        
        # Parse file based on extension
        if filename_lower.endswith('.csv'):
            # Parse CSV
            csv_reader = csv.DictReader(io.StringIO(contents.decode('utf-8')))
            if csv_reader.fieldnames is None:
                raise HTTPException(status_code=400, detail="CSV file is empty or invalid")
            columns = list(csv_reader.fieldnames)
            rows_list = list(csv_reader)
        
        elif filename_lower.endswith('.xlsx'):
            # Parse XLSX
            excel_book = openpyxl.load_workbook(io.BytesIO(contents))
            sheet = excel_book.active
            
            # Get column headers from first row
            columns = [cell.value for cell in sheet[1]]
            if not columns or columns == [None]:
                raise HTTPException(status_code=400, detail="Excel file is empty or has no headers")
            
            # Read rows starting from row 2
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if all(cell is None for cell in row):
                    continue
                row_dict = dict(zip(columns, row))
                rows_list.append(row_dict)
            
            if not rows_list:
                raise HTTPException(status_code=400, detail="Excel file has no data rows")
        
        elif filename_lower.endswith('.xls'):
            # Parse XLS
            excel_book = xlrd.open_workbook(file_contents=contents)
            sheet = excel_book.sheet_by_index(0)
            
            # Get column headers from first row
            if sheet.nrows == 0:
                raise HTTPException(status_code=400, detail="XLS file is empty")
            
            columns = [str(sheet.cell_value(0, col_idx)) for col_idx in range(sheet.ncols)]
            
            # Read rows starting from row 2
            for row_idx in range(1, sheet.nrows):
                row_values = [sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)]
                if all(not val for val in row_values):
                    continue
                row_dict = dict(zip(columns, row_values))
                rows_list.append(row_dict)
            
            if not rows_list:
                raise HTTPException(status_code=400, detail="XLS file has no data rows")
        
        
        # Use proper transaction context manager for SQLAlchemy 2.0
        with engine.begin() as conn:
            # Parse custom column rules if provided
            custom_rules_map = {}
            if column_rules:
                try:
                    custom_rules_map = json.loads(column_rules)
                except:
                    custom_rules_map = {}
            
            # Fetch active rules from database
            rules_result = conn.execute(text("""
                SELECT column_name, rule_type, rule_value
                FROM rules
                WHERE is_active = TRUE
            """))
            
            rules = rules_result.fetchall()
            
            # Build rule_map: {column_name: [{"type": ..., "value": ...}, ...]}
            # Use custom rules if provided for a column, otherwise use system rules
            rule_map = {}
            for r in rules:
                col_name = r[0]
                # If custom rule exists for this column, use it instead, otherwise use system rule
                if col_name in custom_rules_map:
                    custom_rule = custom_rules_map[col_name]
                    if custom_rule and custom_rule.get("type"):
                        rule_map.setdefault(col_name, []).append({
                            "type": custom_rule["type"],
                            "value": custom_rule.get("value", "")
                        })
                else:
                    rule_map.setdefault(col_name, []).append({
                        "type": r[1],
                        "value": r[2]
                    })
            
            # Also add any custom rules for columns that don't have system rules
            for col_name, custom_rule in custom_rules_map.items():
                if col_name not in rule_map and custom_rule and custom_rule.get("type"):
                    rule_map[col_name] = [{
                        "type": custom_rule["type"],
                        "value": custom_rule.get("value", "")
                    }]
            
            # Create a job record with column information
            job_result = conn.execute(text("""
                INSERT INTO jobs (job_name, status, created_at, columns_info)
                VALUES (:job_name, 'processing', NOW(), :columns_info)
                RETURNING id
            """), {
                "job_name": file.filename,
                "columns_info": json.dumps(columns)
            })
            
            job_id = job_result.scalar()
            
            # Process rows
            clean_count = 0
            quarantine_count = 0
            total_count = 0
            row_number = 0
            
            for row in rows_list:
                row_number += 1
                total_count += 1
                row_valid = True
                validation_errors = []
                
                # Apply rules from database
                for column, rules_list in rule_map.items():
                    if column not in row:
                        continue
                        
                    for rule in rules_list:
                        is_valid = apply_rule(
                            row[column],
                            rule["type"],
                            rule["value"]
                        )
                        
                        if not is_valid:
                            row_valid = False
                            validation_errors.append(f"Column '{column}' failed {rule['type']} rule")
                            
                            # Log the validation failure
                            conn.execute(text("""
                                INSERT INTO logs 
                                (job_id, row_number, column_name, original_value, rule_applied, status_color)
                                VALUES (:job_id, :row_number, :column_name, :original_value, :rule_applied, 'red')
                            """), {
                                "job_id": job_id,
                                "row_number": row_number,
                                "column_name": column,
                                "original_value": str(row[column]),
                                "rule_applied": f"{rule['type']}:{rule['value']}"
                            })
                
                # Store all columns as JSON
                row_data = json.dumps(dict(row))
                
                # Insert into appropriate table
                if row_valid:
                    conn.execute(text("""
                        INSERT INTO clean_data (job_id, name, age, row_data, created_at)
                        VALUES (:job_id, :name, :age, :row_data, NOW())
                    """), {
                        "job_id": job_id,
                        "name": row.get("name", ""),
                        "age": int(row.get("age", 0)) if str(row.get("age", "")).isdigit() else 0,
                        "row_data": row_data
                    })
                    clean_count += 1
                    
                    # Log successful validation
                    conn.execute(text("""
                        INSERT INTO logs 
                        (job_id, row_number, status_color)
                        VALUES (:job_id, :row_number, 'green')
                    """), {
                        "job_id": job_id,
                        "row_number": row_number
                    })
                else:
                    conn.execute(text("""
                        INSERT INTO quarantine_data (job_id, name, age, error_reason, row_data, created_at)
                        VALUES (:job_id, :name, :age, :error_reason, :row_data, NOW())
                    """), {
                        "job_id": job_id,
                        "name": row.get("name", ""),
                        "age": int(row.get("age", 0)) if str(row.get("age", "")).isdigit() else 0,
                        "error_reason": "; ".join(validation_errors),
                        "row_data": row_data
                    })
                    quarantine_count += 1
            
            # Update job status
            conn.execute(text("""
                UPDATE jobs
                SET status = 'completed', total_rows = :total, clean_rows = :clean, quarantined_rows = :quarantine
                WHERE id = :job_id
            """), {
                "job_id": job_id,
                "total": total_count,
                "clean": clean_count,
                "quarantine": quarantine_count
            })
            # Transaction commits automatically when exiting the with block
        
        return {
            "message": "File processed successfully",
            "job_id": job_id,
            "total_rows": total_count,
            "clean_rows": clean_count,
            "quarantined_rows": quarantine_count,
            "status": "completed"
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/jobs/{job_id}")
def get_job_status(job_id: int):
    """Get status of a specific job"""
    try:
        conn = engine.connect()
        result = conn.execute(text("""
            SELECT id, job_name, status, total_rows, clean_rows, quarantined_rows, created_at
            FROM jobs
            WHERE id = :job_id
        """), {"job_id": job_id})
        
        job = result.fetchone()
        conn.close()
        
        if not job:
            raise HTTPException(status_code=404, detail="Job not found")
        
        return {
            "id": job[0],
            "job_name": job[1],
            "status": job[2],
            "total_rows": job[3],
            "clean_rows": job[4],
            "quarantined_rows": job[5],
            "created_at": job[6]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/rules")
def get_rules():
    """Get all active rules"""
    try:
        with engine.connect() as conn:
            result = conn.execute(text("""
                SELECT id, column_name, rule_type, rule_value, is_active
                FROM rules
                WHERE is_active = TRUE
                ORDER BY id DESC
            """))
            
            rules = result.fetchall()
            
            return [
                {
                    "id": r[0],
                    "column_name": r[1],
                    "rule_type": r[2],
                    "rule_value": r[3],
                    "is_active": r[4]
                }
                for r in rules
            ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/add-rule")
def add_rule(column_name: str, rule_type: str, rule_value: str):
    """Add a new validation rule"""
    try:
        with engine.connect() as conn:
            conn.execute(text("""
                INSERT INTO rules (column_name, rule_type, rule_value, is_active)
                VALUES (:column, :type, :value, TRUE)
            """), {
                "column": column_name,
                "type": rule_type,
                "value": rule_value
            })
            conn.commit()
        return {"message": "Rule added successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/rules/{rule_id}")
def update_rule(rule_id: int, column_name: str, rule_type: str, rule_value: str):
    """Update an existing validation rule"""
    try:
        with engine.connect() as conn:
            conn.execute(text("""
                UPDATE rules
                SET column_name = :column, rule_type = :type, rule_value = :value
                WHERE id = :id
            """), {
                "column": column_name,
                "type": rule_type,
                "value": rule_value,
                "id": rule_id
            })
            conn.commit()
        return {"message": "Rule updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/jobs")
def get_jobs():
    """Get all jobs sorted by most recent first"""
    try:
        with engine.connect() as conn:
            result = conn.execute(text("""
                SELECT id, job_name, status, total_rows, clean_rows, quarantined_rows, created_at
                FROM jobs
                ORDER BY created_at DESC
                LIMIT 100
            """))
            jobs = result.fetchall()
        
        return [
            {
                "id": j[0],
                "job_name": j[1],
                "status": j[2],
                "total_rows": j[3],
                "clean_rows": j[4],
                "quarantined_rows": j[5],
                "created_at": str(j[6])
            }
            for j in jobs
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/quarantine")
def get_quarantine():
    """Get all quarantined rows"""
    try:
        with engine.connect() as conn:
            result = conn.execute(text("""
                SELECT id, job_id, name, age, error_reason, created_at
                FROM quarantine_data
                ORDER BY created_at DESC
            """))
            rows = result.fetchall()
        
        return [
            {
                "id": r[0],
                "job_id": r[1],
                "name": r[2],
                "age": r[3],
                "error_reason": r[4],
                "created_at": str(r[5])
            }
            for r in rows
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/update-quarantine/{row_id}")
def update_quarantine(row_id: int, name: str, age: int):
    """Update a quarantined row"""
    try:
        with engine.connect() as conn:
            conn.execute(text("""
                UPDATE quarantine_data
                SET name = :name, age = :age
                WHERE id = :id
            """), {
                "name": name,
                "age": age,
                "id": row_id
            })
            conn.commit()
        return {"message": "Row updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/revalidate/{row_id}")
def revalidate_row(row_id: int):
    """Re-validate a quarantined row and move to clean_data if valid"""
    try:
        with engine.connect() as conn:
            # Fetch quarantined row
            row_result = conn.execute(text("""
                SELECT id, job_id, name, age, error_reason
                FROM quarantine_data
                WHERE id = :id
            """), {"id": row_id})
            
            row = row_result.fetchone()
            
            if not row:
                raise HTTPException(status_code=404, detail="Row not found")
            
            # Fetch active rules
            rules_result = conn.execute(text("""
                SELECT column_name, rule_type, rule_value
                FROM rules
                WHERE is_active = TRUE
            """))
            
            rules = rules_result.fetchall()
            
            # Build rule_map
            rule_map = {}
            for r in rules:
                rule_map.setdefault(r[0], []).append({
                    "type": r[1],
                    "value": r[2]
                })
            
            # Re-validate the row
            row_valid = True
            validation_errors = []
            
            # Check name
            if "name" in rule_map:
                for rule in rule_map["name"]:
                    if not apply_rule(row[2], rule["type"], rule["value"]):
                        row_valid = False
                        validation_errors.append(f"name failed {rule['type']}")
            
            # Check age
            if "age" in rule_map:
                for rule in rule_map["age"]:
                    if not apply_rule(row[3], rule["type"], rule["value"]):
                        row_valid = False
                        validation_errors.append(f"age failed {rule['type']}")
            
            if row_valid:
                # Move to clean_data
                conn.execute(text("""
                    INSERT INTO clean_data (job_id, name, age, created_at)
                    VALUES (:job_id, :name, :age, NOW())
                """), {
                    "job_id": row[1],
                    "name": row[2],
                    "age": row[3]
                })
                
                # Delete from quarantine
                conn.execute(text("""
                    DELETE FROM quarantine_data
                    WHERE id = :id
                """), {"id": row_id})
                
                # Log the correction
                conn.execute(text("""
                    INSERT INTO logs (job_id, row_number, column_name, original_value, final_value, status_color, rule_applied)
                    VALUES (:job_id, :row_number, 'system', :original, :final, 'green', 'revalidation')
                """), {
                    "job_id": row[1],
                    "row_number": row[0],
                    "original": row[4],
                    "final": "CORRECTED"
                })
                
                conn.commit()
                return {"status": "success", "message": "Row moved to clean_data"}
            else:
                return {"status": "invalid", "message": "Row still invalid", "errors": validation_errors}
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/logs/{job_id}")
def get_logs(job_id: int):
    """Get all logs for a specific job"""
    try:
        with engine.connect() as conn:
            result = conn.execute(text("""
                SELECT id, job_id, row_number, column_name, original_value, final_value, status_color, rule_applied, created_at
                FROM logs
                WHERE job_id = :job_id
                ORDER BY row_number, id
            """), {"job_id": job_id})
            
            logs = result.fetchall()
        
        return [
            {
                "id": l[0],
                "job_id": l[1],
                "row_number": l[2],
                "column_name": l[3] or "",
                "original_value": l[4] or "",
                "final_value": l[5] or "",
                "status_color": l[6] or "unknown",
                "rule_applied": l[7] or "",
                "created_at": str(l[8])
            }
            for l in logs
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/revalidate-job/{job_id}")
def revalidate_job(job_id: int):
    """Revalidate a job with current system rules"""
    try:
        with engine.begin() as conn:
            # Get all data (clean and quarantined) for this job
            all_data = conn.execute(text("""
                SELECT row_data FROM clean_data WHERE job_id = :job_id
                UNION ALL
                SELECT row_data FROM quarantine_data WHERE job_id = :job_id
            """), {"job_id": job_id}).fetchall()
            
            if not all_data:
                raise HTTPException(status_code=404, detail="No data found for this job")
            
            # Get current system rules
            rules_result = conn.execute(text("""
                SELECT column_name, rule_type, rule_value
                FROM rules
                WHERE is_active = TRUE
            """))
            
            rules = {}
            for r in rules_result.fetchall():
                col_name = r[0]
                if col_name not in rules:
                    rules[col_name] = []
                rules[col_name].append({"type": r[1], "value": r[2]})
            
            # Clear old logs for this job
            conn.execute(text("DELETE FROM logs WHERE job_id = :job_id"), {"job_id": job_id})
            
            # Clear old data tables
            conn.execute(text("DELETE FROM clean_data WHERE job_id = :job_id"), {"job_id": job_id})
            conn.execute(text("DELETE FROM quarantine_data WHERE job_id = :job_id"), {"job_id": job_id})
            
            # Revalidate all rows
            clean_count = 0
            quarantine_count = 0
            row_number = 0
            
            for row_data_json in all_data:
                row_number += 1
                import json
                row = json.loads(row_data_json[0]) if isinstance(row_data_json[0], str) else row_data_json[0]
                
                row_valid = True
                validation_errors = []
                
                # Apply current rules
                for column, rules_list in rules.items():
                    if column not in row:
                        continue
                    
                    for rule in rules_list:
                        is_valid = apply_rule(row[column], rule["type"], rule["value"])
                        
                        if not is_valid:
                            row_valid = False
                            validation_errors.append(f"Column '{column}' failed {rule['type']} rule")
                            conn.execute(text("""
                                INSERT INTO logs 
                                (job_id, row_number, column_name, original_value, rule_applied, status_color)
                                VALUES (:job_id, :row_number, :column_name, :original_value, :rule_applied, 'red')
                            """), {
                                "job_id": job_id,
                                "row_number": row_number,
                                "column_name": column,
                                "original_value": str(row[column]),
                                "rule_applied": f"{rule['type']}: {rule['value']}"
                            })
                
                # Store result
                if row_valid:
                    clean_count += 1
                    conn.execute(text("""
                        INSERT INTO clean_data (job_id, row_number, name, age, row_data)
                        VALUES (:job_id, :row_number, :name, :age, :row_data)
                    """), {
                        "job_id": job_id,
                        "row_number": row_number,
                        "name": row.get("name", ""),
                        "age": row.get("age", ""),
                        "row_data": json.dumps(row)
                    })
                else:
                    quarantine_count += 1
                    conn.execute(text("""
                        INSERT INTO quarantine_data (job_id, row_number, name, age, row_data, validation_errors)
                        VALUES (:job_id, :row_number, :name, :age, :row_data, :errors)
                    """), {
                        "job_id": job_id,
                        "row_number": row_number,
                        "name": row.get("name", ""),
                        "age": row.get("age", ""),
                        "row_data": json.dumps(row),
                        "errors": ", ".join(validation_errors)
                    })
            
            # Update job counts
            conn.execute(text("""
                UPDATE jobs
                SET clean_rows = :clean, quarantined_rows = :quarantine
                WHERE id = :job_id
            """), {
                "clean": clean_count,
                "quarantine": quarantine_count,
                "job_id": job_id
            })
        
        return {
            "message": "Job revalidated successfully",
            "job_id": job_id,
            "clean_rows": clean_count,
            "quarantined_rows": quarantine_count,
            "total_rows": clean_count + quarantine_count
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/clean-data")
def get_clean_data(limit: int = 5, job_id: int = None):
    """Get clean data from the database with dynamic columns"""
    try:
        import json
        with engine.connect() as conn:
            if job_id:
                # Get data from specific job
                result = conn.execute(text(f"""
                    SELECT id, job_id, name, age, row_data, created_at
                    FROM clean_data
                    WHERE job_id = :job_id
                    ORDER BY id ASC
                    LIMIT {limit}
                """), {"job_id": job_id})
            else:
                # Get data from most recent job that has clean data
                result = conn.execute(text(f"""
                    SELECT id, job_id, name, age, row_data, created_at
                    FROM clean_data
                    WHERE row_data IS NOT NULL
                    ORDER BY job_id DESC, id ASC
                    LIMIT {limit}
                """))
            
            rows = result.fetchall()
        
        data = []
        for r in rows:
            row_dict = {
                "id": r[0],
                "job_id": r[1],
                "created_at": str(r[5]) if r[5] else None
            }
            
            # Add row_data fields
            if r[4]:  # row_data column
                try:
                    row_data = json.loads(r[4])
                    row_dict.update(row_data)
                except:
                    row_dict["name"] = r[2] or ""
                    row_dict["age"] = r[3]
            else:
                # Fallback to structured columns
                row_dict["name"] = r[2] or ""
                row_dict["age"] = r[3]
            
            data.append(row_dict)
        
        return data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/job-columns/{job_id}")
def get_job_columns(job_id: int):
    """Get the column names for a specific job"""
    try:
        import json
        with engine.connect() as conn:
            result = conn.execute(text("""
                SELECT columns_info
                FROM jobs
                WHERE id = :job_id
            """), {"job_id": job_id})
            
            job = result.fetchone()
            
            if not job or not job[0]:
                # Fallback: try to infer columns from clean_data
                return ["id", "job_id", "name", "age", "created_at"]
            
            columns = json.loads(job[0])
            return columns
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

